# This Python file uses the following encoding: utf-8

from aula.apps.extPreinscripcio.models import Preinscripcio

from aula.apps.missatgeria.missatges_a_usuaris import tipusMissatge, IMPORTACIO_PREINSCRIPCIO_FINALITZADA
from aula.apps.missatgeria.models import Missatge

from openpyxl import load_workbook
from django.contrib.auth.models import Group
import json

def assignaDades(preinscripcio, index, value, colnames, col_indexs):
    if index in col_indexs and value:
        field=colnames[col_indexs[index]]
        if isinstance(field, dict):
            if 'case' in field and field['case']:
                value=str(value).lower() if field['case']=='L' else str(value).upper()
            if 'append' in field and field['append']:
                if field['field'] not in preinscripcio:
                    preinscripcio[field['field']]=value
                else:
                    preinscripcio[field['field']]=str(preinscripcio[field['field']])+" "+str(value)
            else:
                if 'field' in field and field['field']:
                    preinscripcio[field['field']]=value
        else:
            preinscripcio[field]=value
    return preinscripcio

def sincronitza(f, user = None):
    
    errors = []
    warnings= []
    infos= []
    
    try:
        # Carregar full de càlcul
        wb2 = load_workbook(f)
        if len(wb2.worksheets)!=1:
            # Si té més d'una pestanya --> error
            errors.append('Fitxer incorrecte sheets')
            return {'errors': errors, 'warnings': [], 'infos': []}
    except:
        errors.append('Fitxer incorrecte')
        return {'errors': errors, 'warnings': [], 'infos': []}
    
    info_nAlumnesLlegits=0
    
    # Carregar full de càlcul
    full = wb2.active
    max_row = full.max_row

    # columnes que s'importaran,  camp excel : camp base de dades 
    colnames = {
        'estat sol·licitud':{},
        'nom': 'nom',
        'primer cognom': 'cognoms',
        'segon cognom': {'field': 'cognoms', 'append' : True},
        'identificació ralc': 'ralc',
        'codi ensenyament p1': 'codiestudis',
        'nom ensenyament p1': 'nomestudis',
        'codi modalitat': 'codimodalitat',
        'modalitat': 'nommodalitat',
        'curs p1': 'curs',
        'règim p1': 'regim',
        'torn p1': 'torn',
        'dni': {'field': 'identificador', 'case' : 'U'},
        'nie': {'field': 'identificador', 'case' : 'U'},
        'pass': {'field': 'identificador', 'case' : 'U'},
        'tis': 'tis',
        'data naixement': 'naixement',
        'sexe': 'sexe',
        'nacionalitat': 'nacionalitat',
        'país naixement': 'paisnaixement',
        'tipus via': 'adreça',
        'nom via': {'field': 'adreça', 'append' : True},
        'número via': {'field': 'adreça', 'append' : True},
        'altres dades': {'field': 'adreça', 'append' : True},
        'província residència': 'provincia',
        'municipi residència': 'municipi',
        'localitat residència': 'localitat',
        'cp': 'cp',
        'país residència': 'paisresidencia',
        'telèfon': 'telefon',
        'correu electrònic': {'field': 'correu', 'case' : 'L'},
        'tipus doc. tutor 1': 'tdoctut1',
        'núm. doc. tutor 1': 'doctut1',
        'nom tutor 1': 'nomtut1',
        'primer cognom tutor 1': 'cognomstut1',
        'segon cognom tutor 1': {'field': 'cognomstut1', 'append' : True},
        'tipus doc. tutor 2': 'tdoctut2',
        'núm. doc. tutor 2': 'doctut2',
        'nom tutor 2': 'nomtut2',
        'primer cognom tutor 2': 'cognomstut2',
        'segon cognom tutor 2': {'field': 'cognomstut2', 'append' : True},
        'codi centre proc.': 'codicentreprocedencia',
        'nom centre proc.': 'centreprocedencia',
        'codi ensenyament proc.': 'codiestudisprocedencia',
        'nom ensenyament proc.': 'estudisprocedencia',
        'curs proc.': 'cursestudisprocedencia',
        }
    
    rows = list(wb2.active.rows)
    col_indexs = {n: str(cell.value).lower() for n, cell in enumerate(rows[0])
                   if str(cell.value).lower() in colnames}
    
    for row in rows[1:max_row]:
       
        preinscripcio={}
        estat=None
        for index, cell in enumerate(row):
            if bool(cell) and bool(cell.value) and isinstance(cell.value, str):
                cell.value=cell.value.strip()
            preinscripcio=assignaDades(preinscripcio, index, cell.value, colnames, col_indexs)
            if index in col_indexs and col_indexs[index]=='estat sol·licitud':
                estat=cell.value

        if 'cp' in preinscripcio and preinscripcio['cp'].startswith("="): preinscripcio['cp']=preinscripcio['cp'][2:7]

        if estat=='Validada':
            try:
                p=Preinscripcio(**preinscripcio)
                query=Preinscripcio.objects.filter(ralc=p.ralc)
                if query:
                    p.pk=query[0].pk
                p.save()
                info_nAlumnesLlegits += 1
            except Exception as e:
                errors.append(str(e)+": "+json.dumps(preinscripcio))
                

    infos.append(u'{0} alumnes llegits'.format(info_nAlumnesLlegits) )

    missatge = IMPORTACIO_PREINSCRIPCIO_FINALITZADA
    tipus_de_missatge = tipusMissatge(missatge)
    msg = Missatge(
                remitent= user,
                text_missatge = missatge,
                tipus_de_missatge = tipus_de_missatge)
    msg.afegeix_errors( errors )
    msg.afegeix_warnings(warnings)
    msg.afegeix_infos(infos)
    importancia = 'VI' if len( errors )> 0 else 'IN'
    grupDireccio =  Group.objects.get( name = 'direcció' )
    msg.envia_a_grup( grupDireccio , importancia=importancia)

    return { 'errors': errors, 'warnings': warnings, 'infos': infos }
